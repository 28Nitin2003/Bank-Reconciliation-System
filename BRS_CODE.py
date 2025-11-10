import streamlit as st
import os, re, warnings
import pandas as pd
from io import BytesIO
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")

# ----------------------------------------------------
# Streamlit Page Setup + Styling
# ----------------------------------------------------
st.set_page_config(page_title="Bank & SAP Data Checker", page_icon="🏦", layout="wide")

st.markdown("""
<style>
html, body, .stApp {
    background:#f8fafc;
    font-family:"Segoe UI",sans-serif;
    color:#1e293b;
}

/* Header */
.app-header {
    background:linear-gradient(90deg,#06b6d4,#0891b2);
    padding:26px;
    border-radius:14px;
    text-align:center;
    box-shadow:0 3px 10px rgba(0,0,0,0.1);
}
.app-header h2 {color:#fff;margin:0;font-weight:600;}
.app-header p {color:#e0f7fa;margin-top:6px;font-size:15px;}

/* Sidebar */
section[data-testid="stSidebar"] {
    background:linear-gradient(180deg,#0f172a 0%,#1e293b 100%);
    box-shadow:2px 0 8px rgba(0,0,0,0.25);
}
section[data-testid="stSidebar"] h1,
section[data-testid="stSidebar"] h2,
section[data-testid="stSidebar"] label,
section[data-testid="stSidebar"] p {color:#f8fafc !important;}
section[data-testid="stSidebar"] .stMarkdown {color:#f1f5f9 !important;}

/* Dark File Uploader Box */
[data-testid="stFileUploader"] {
    background-color:#1e293b;
    border:1px solid #334155;
    border-radius:10px;
    padding:16px;
    color:#f8fafc !important;
    box-shadow:inset 0 0 6px rgba(0,0,0,0.5);
}
[data-testid="stFileUploader"] label div,
[data-testid="stFileUploader"] > div > div > div > span {
    color:#f8fafc !important;
    font-weight:600;
}

/* Buttons */
div.stButton > button, .stDownloadButton>button {
    background:linear-gradient(90deg,#06b6d4,#0891b2);
    color:white;font-weight:600;border:none;
    border-radius:8px;padding:10px 22px;
    transition:all .3s ease;font-size:15px;
}
div.stButton > button:hover, .stDownloadButton>button:hover {
    background:linear-gradient(90deg,#0e7490,#06b6d4);
    transform:translateY(-2px);
}

/* Metrics */
[data-testid="stMetricValue"] {color:#0891b2 !important;font-weight:700;}
[data-testid="stMetricLabel"] {color:#475569 !important;}

/* DataFrame */
.stDataFrame {
    border:1px solid #e2e8f0;
    border-radius:10px;
    background:white;
    box-shadow:0 1px 6px rgba(0,0,0,0.06);
}

/* Legend */
.legend-box {
    background:#e0f2f1;
    padding:10px 15px;
    border-radius:8px;
    margin-top:10px;
    color:#134e4a;
    font-size:14px;
}
</style>
""", unsafe_allow_html=True)

# ----------------------------------------------------
# Header
# ----------------------------------------------------
st.markdown("""
<div class='app-header'>
  <h2>🏦 Bank & SAP Data Checker</h2>
  <p>Compare SAP and Bank data — generate clean Excel reconciliation reports instantly</p>
</div>
""", unsafe_allow_html=True)

# ----------------------------------------------------
# Sidebar Upload Section
# ----------------------------------------------------
st.sidebar.header("📂 Upload Your Files")
bank_file = st.sidebar.file_uploader("🏦 Upload Bank Excel File", type=["xlsx","xls","csv"])
bank_sheet = st.sidebar.text_input("📄 Bank File Sheet Name (optional)")
sap_file  = st.sidebar.file_uploader("💼 Upload SAP Excel File",  type=["xlsx","xls","csv"])

st.sidebar.markdown("---")
st.sidebar.markdown("### ⚙️ Matching Settings")
fuzzy_threshold = st.sidebar.slider("Fuzzy Match Threshold (%)",50,100,60,5)

os.makedirs("uploaded_files",exist_ok=True)
sap_path=bank_path=None
if sap_file:
    sap_path="uploaded_files/SAP.xlsx"
    with open(sap_path,"wb") as f:f.write(sap_file.getbuffer())
    st.sidebar.success("✅ SAP File Uploaded")
if bank_file:
    bank_path="uploaded_files/Bank.xlsx"
    with open(bank_path,"wb") as f:f.write(bank_file.getbuffer())
    st.sidebar.success("✅ Bank File Uploaded")

# ----------------------------------------------------
# Account Type
# ----------------------------------------------------
st.markdown("### 🏦 Select Account Type")
acct_type=st.selectbox("Account Type",["G/L Account","BRS Account"],index=1)

# ----------------------------------------------------
# Processor
# ----------------------------------------------------
class Processor:
    def __init__(self,b,s,f,a):
        self.b,self.s,self.f,self.a=b,s,f,a

    def _prep(self,df):
        header_idx=None
        for i in range(len(df)):
            vals=df.iloc[i,0:4].astype(str).str.lower()
            if any(v in["date","txn date","transaction date"] for v in vals):
                header_idx=i;break
        if header_idx is None: return pd.DataFrame()
        df.columns=df.iloc[header_idx]
        df=df[header_idx+1:]
        df.columns=[str(c).title() for c in df.columns]
        for w in["Withdrawal","Withdrawals","Debit","Dr Amount"]:
            if w in df.columns:
                df.rename(columns={w:"Withdrawals"},inplace=True)
        df=df[df["Withdrawals"].notna()]
        return df

    def load_bank(self,sh=None):
        if not self.b: return False
        data=pd.read_excel(self.b,header=None,sheet_name=sh or None)
        if isinstance(data,dict):
            frames=[self._prep(d) for d in data.values()]
            frames=[f for f in frames if not f.empty]
            self.df=pd.concat(frames,ignore_index=True)
        else:
            self.df=self._prep(data)
        return True

    def load_sap(self):
        self.df2=pd.read_excel(self.s)
        col="Amount in LC" if self.a=="BRS Account" else "Amount in Local Currency"
        if col not in self.df2.columns:
            st.error(f"❌ Missing column '{col}' in SAP file.");return False
        self.df2[col]=pd.to_numeric(self.df2[col],errors="coerce")
        return True

    def match(self):
        bank=self.df.copy(); sap=self.df2.copy()
        col="Amount in LC" if self.a=="BRS Account" else "Amount in Local Currency"
        bank["Withdrawals"]=pd.to_numeric(bank["Withdrawals"],errors="coerce")
        sap["status"]="Not Found in Bank Statement"
        for i,row in sap.iterrows():
            amt=row[col]
            if pd.isna(amt): continue
            match_rows=bank[bank["Withdrawals"]==amt]
            if len(match_rows)==1:
                sap.at[i,"status"]="100% Matched"
            elif len(match_rows)>1:
                sap.at[i,"status"]="Multiple Matches"

        # entries present in Bank but not in SAP
        sap_amts=set(sap[col].dropna())
        bank_amts=set(bank["Withdrawals"].dropna())
        extra_bank=bank_amts-sap_amts
        if len(extra_bank)>0:
            extra=pd.DataFrame({col:list(extra_bank)})
            extra["status"]="Not Found in SAP Record"
            sap=pd.concat([sap,extra],ignore_index=True)

        # Remove incomplete rows
        non_empty_cols=sap.drop(columns=["status",col],errors="ignore").count(axis=1)
        sap=sap[(sap["status"]!="Not Found in SAP Record") | (non_empty_cols>0)]
        self.final=sap

    def excel(self):
        buf=BytesIO()
        with pd.ExcelWriter(buf,engine="openpyxl") as writer:
            self.final.to_excel(writer,index=False,sheet_name="Data")
            wb=writer.book
            ws=wb["Data"]

            green=PatternFill(start_color="90EE90",end_color="90EE90",fill_type="solid")
            orange=PatternFill(start_color="FFB347",end_color="FFB347",fill_type="solid")
            red=PatternFill(start_color="FF9999",end_color="FF9999",fill_type="solid")
            yellow=PatternFill(start_color="FFF2CC",end_color="FFF2CC",fill_type="solid")

            status_col=None
            for idx,cell in enumerate(ws[1],start=1):
                if str(cell.value).lower()=="status": status_col=idx;break

            if status_col:
                for r in range(2,ws.max_row+1):
                    val=str(ws.cell(row=r,column=status_col).value or "").lower()
                    if "100%" in val: ws.cell(row=r,column=status_col).fill=green
                    elif "multiple" in val: ws.cell(row=r,column=status_col).fill=orange
                    elif "bank statement" in val: ws.cell(row=r,column=status_col).fill=red
                    elif "sap record" in val: ws.cell(row=r,column=status_col).fill=yellow

            for col_cells in ws.columns:
                max_len=max(len(str(c.value)) if c.value else 0 for c in col_cells)
                ws.column_dimensions[get_column_letter(col_cells[0].column)].width=max_len+2

            total=len(self.final)
            matched=(self.final["status"]=="100% Matched").sum()
            multi=(self.final["status"]=="Multiple Matches").sum()
            not_bank=(self.final["status"].str.contains("Bank",case=False)).sum()
            not_sap=(self.final["status"].str.contains("SAP",case=False)).sum()

            summary=wb.create_sheet("Summary")
            summary["A1"]="BANK RECONCILIATION SUMMARY"
            summary["A1"].font=Font(bold=True,size=14,color="003366")
            summary.merge_cells("A1:D1")
            summary["A1"].alignment=Alignment(horizontal="center")

            rows=[
                f"Bank File: {os.path.basename(self.b)}",
                f"SAP File:  {os.path.basename(self.s)}",
                "",
                f"Total Records: {total}",
                f"Matched: {matched}",
                f"Multiple Matches: {multi}",
                f"Not Found in Bank Statement: {not_bank}",
                f"Not Found in SAP Record: {not_sap}",
                "",
                f"{not_bank} records found in SAP but missing in Bank Statement.",
                f"{not_sap} records found in Bank but missing in SAP."
            ]
            for i,txt in enumerate(rows,start=3):
                summary[f"A{i}"]=txt
                if i>=11: summary[f"A{i}"].font=Font(italic=True,color="555555")

            summary.column_dimensions["A"].width=70

        buf.seek(0)
        return buf

# ----------------------------------------------------
# Processing Section
# ----------------------------------------------------
st.markdown("### 🚀 Process Files")

if st.button("Generate Final Statement"):
    if not sap_file or not bank_file:
        st.error("❌ Please upload both SAP and Bank files.")
    else:
        with st.spinner("Processing your files..."):
            p=Processor(bank_path,sap_path,fuzzy_threshold,acct_type)
            if p.load_bank(bank_sheet) and p.load_sap():
                p.match()
                excel_data=p.excel()

                st.success("✅ Final statement generated successfully!")
                # ✅ KPI Summary
                matched=(p.final["status"]=="100% Matched").sum()
                multiple=(p.final["status"]=="Multiple Matches").sum()
                notfound=(p.final["status"].str.contains("Not",case=False)).sum()

                c1,c2,c3=st.columns(3)
                c1.metric("✅ 100% Matched",matched)
                c2.metric("🟠 Multiple Matches",multiple)
                c3.metric("❌ Not Found",notfound)

                st.markdown("### 📊 Preview (First 10 Rows)")
                st.dataframe(p.final.head(10))

                st.download_button(
                    label="⬇️ Download Final Statement (Color Coded)",
                    data=excel_data,
                    file_name=f"{'BRS' if acct_type=='BRS Account' else 'GL'}_Match_Status.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

                st.markdown("""
                <div class="legend-box">
                <b>Legend:</b> 🟩 Matched (Green) &nbsp;&nbsp; 🟧 Multiple Matches (Orange) &nbsp;&nbsp; 🟥 Not Found in Bank Statement (Red) &nbsp;&nbsp; 🟨 Not Found in SAP Record (Yellow)
                </div>
                """, unsafe_allow_html=True)
